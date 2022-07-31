from openpyxl import load_workbook
 
#load excel file
workbook = load_workbook(filename="d:/GitHub/testrepo/python/csalad01.xlsx")
 
  
 
#open workbook
sheet = workbook.active
 
#modify the desired cell
# sheet["A1"] = "Name"
# sheet["B1"] = "Name"
# sheet["C1"] = "Name"
# sheet["D1"] = "Name"

# ws.merge_cells('B2:F4')
# top_left_cell = ws['B2']
# top_left_cell.value = "My Cell"
 
#save the file
workbook.save(filename="D:/GitHub/testrepo/python/csalad01.xlsx")